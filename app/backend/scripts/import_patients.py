"""
Bulk import patients from CSV or XLSX file.

Usage:
  python scripts/import_patients.py --file data/patients_100.xlsx
  python scripts/import_patients.py --file data/patients_template.csv --dry-run

Encoding: UTF-8 with BOM recommended for CSV files with Vietnamese text.
"""
import argparse
import csv
import datetime
import uuid
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

ALLOWED_STATUS = {"active", "discharged", "deceased"}
UUID_PREFIX = "20000000-0000-0000-0000-"


def parse_args():
    p = argparse.ArgumentParser(description="Bulk import patients from CSV/XLSX")
    p.add_argument("--file", required=True, help="Path to CSV or XLSX file")
    p.add_argument("--dry-run", action="store_true", help="Validate only, no insert")
    return p.parse_args()


def read_csv(path: Path):
    rows = []
    with open(path, "r", encoding="utf-8-sig") as f:
        lines = [line for line in f if not line.startswith("#")]
    reader = csv.DictReader(lines)
    for row in reader:
        rows.append(row)
    return rows


def read_xlsx(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Install openpyxl: pip install openpyxl")
        raise
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))  # noqa: B905
    wb.close()
    return rows


def validate_row(row: dict, index: int):
    errors = []
    mrn = (row.get("mrn") or "").strip()
    name = (row.get("full_name") or "").strip()
    dob_str = (row.get("dob") or "").strip()
    status = (row.get("status") or "active").strip().lower()

    if not mrn:
        errors.append("mrn is required")
    if not name:
        errors.append("full_name is required")
    if dob_str:
        try:
            datetime.date.fromisoformat(dob_str)
        except ValueError:
            errors.append(f"Invalid dob format: {dob_str} (expected YYYY-MM-DD)")
    if status not in ALLOWED_STATUS:
        errors.append(f"Invalid status: {status} (allowed: {sorted(ALLOWED_STATUS)})")

    return errors


async def import_patients(file_path: str, dry_run: bool = False):
    path = Path(file_path)
    if not path.exists():
        print(f"ERROR: File not found: {path}")
        return

    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = read_csv(path)
    elif suffix in (".xlsx", ".xls"):
        rows = read_xlsx(path)
    else:
        print(f"ERROR: Unsupported format: {suffix} (use .csv or .xlsx)")
        return

    print(f"Read {len(rows)} rows from {path.name}")

    # Validate
    errors = []
    valid = []
    for i, row in enumerate(rows):
        row_errors = validate_row(row, i + 1)
        if row_errors:
            errors.append((i + 1, row.get("mrn", "?"), row_errors))
        else:
            valid.append(row)

    if errors:
        print(f"\nValidation errors ({len(errors)}):")
        for line, mrn, errs in errors:
            print(f"  Row {line} (MRN={mrn}): {', '.join(errs)}")
        if dry_run:
            return

    print(f"Valid rows: {len(valid)}")

    if dry_run:
        print("[DRY RUN] No data written. Preview of first 5:")
        for r in valid[:5]:
            safe_name = r['full_name'].encode('ascii', 'replace').decode()
            safe_dept = (r.get('department') or '').encode('ascii', 'replace').decode()
            print(f"  {r['mrn']} | {safe_name} | {r.get('dob','')} | {safe_dept} | {r.get('status','active')}")
        return

    # Import
    from hospital_ai.db.session import get_session
    from hospital_ai.db.models import Patient
    from sqlalchemy import select

    imported = 0
    skipped = 0
    failed = 0

    async for session in get_session():
        for i, row in enumerate(valid):
            mrn = row["mrn"].strip()
            name = row["full_name"].strip()
            dob_str = (row.get("dob") or "").strip()
            department = (row.get("department") or "").strip()
            status = (row.get("status") or "active").strip().lower()

            existing = await session.execute(
                select(Patient).where(Patient.mrn == mrn)
            )
            if existing.scalar_one_or_none():
                skipped += 1
                continue

            try:
                patient_id = uuid.UUID(f"{UUID_PREFIX}{i + 6:012d}")
                dob = datetime.date.fromisoformat(dob_str) if dob_str else None
                patient = Patient(
                    id=patient_id,
                    mrn=mrn,
                    full_name=name,
                    dob=dob,
                    department=department or None,
                    status=status,
                )
                session.add(patient)
                imported += 1
            except Exception as exc:
                print(f"  ERROR row {i+1} (MRN={mrn}): {exc}")
                failed += 1

        if imported > 0:
            await session.commit()
            print(f"\nCommitted {imported} patients.")
        break  # single session from generator

    print(f"\nResults: Imported={imported}, Skipped (exists)={skipped}, Failed={failed}")
    if imported > 0:
        mrn_range = f"{valid[0]['mrn']} – {valid[-1]['mrn']}"
        print(f"MRN range: {mrn_range}")


if __name__ == "__main__":
    args = parse_args()
    import asyncio
    asyncio.run(import_patients(args.file, args.dry_run))
