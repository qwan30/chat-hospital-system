"""Excel document loader using openpyxl."""

from __future__ import annotations

from pathlib import Path

from hospital_ai.core.errors import ExternalServiceError
from hospital_ai.services.loaders.base import BaseDocumentLoader, LoadedPage


class ExcelLoader(BaseDocumentLoader):
    """Extract text from Excel files.

    Each sheet becomes a separate LoadedPage. Data is converted to pipe-delimited
    text format suitable for embedding and retrieval.
    """

    def supported_extensions(self) -> set[str]:
        return {".xlsx", ".xls"}

    def load(self, file_path: Path, mime_type: str = "") -> list[LoadedPage]:
        if not file_path.exists():
            raise ExternalServiceError(f"Excel file not found: {file_path}")

        try:
            from openpyxl import load_workbook  # type: ignore[import-untyped]
        except ImportError:
            raise ExternalServiceError("openpyxl is not installed. Install it with `pip install openpyxl`.") from None

        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as exc:
            raise ExternalServiceError(f"Failed to open Excel file: {exc}") from exc

        pages: list[LoadedPage] = []
        for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cell_values = [str(cell) if cell is not None else "" for cell in row]
                row_text = " | ".join(cell_values).strip()
                if row_text.replace("|", "").strip():
                    rows.append(row_text)

            if rows:
                text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
                pages.append(
                    LoadedPage(
                        page_number=sheet_idx,
                        text=text,
                        confidence=1.0,
                        metadata={"sheet_name": sheet_name},
                    )
                )
        wb.close()

        if not pages:
            raise ExternalServiceError("Excel file contains no data.")
        return pages
